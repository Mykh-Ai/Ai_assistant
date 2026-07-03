from __future__ import annotations

import asyncio
import json

from datetime import datetime, date
from pathlib import Path

from openpyxl import load_workbook

from bot.services.db import init_db
from bot.services.work_time import (
    WorkTimeCandidate,
    WorkTimeService,
    format_candidate_preview,
    format_day_summary,
    parse_close_candidate,
    parse_duration_entry_candidate,
    parse_lunch_break_minutes,
    parse_explicit_month,
    parse_manual_range_candidate,
    parse_report_month,
    resolve_work_time_entry_candidate,
    _format_duration,
)


class _FakeChoice:
    def __init__(self, content: str) -> None:
        self.message = type('_Message', (), {'content': content})()


class _FakeResponse:
    def __init__(self, content: str) -> None:
        self.choices = [_FakeChoice(content)]


class _WorkTimeSlotOpenAIFake:
    output = '{"canonical":"unknown"}'
    last_payload: dict | None = None

    def __init__(self, *, api_key: str) -> None:
        self.api_key = api_key
        self.chat = type('_Chat', (), {'completions': self})()

    async def create(self, **kwargs):
        _WorkTimeSlotOpenAIFake.last_payload = json.loads(kwargs['messages'][1]['content'])
        return _FakeResponse(_WorkTimeSlotOpenAIFake.output)


def test_close_with_duration_calculates_end_time_and_total(tmp_path: Path) -> None:
    db_path = tmp_path / 'test.db'
    init_db(db_path)
    service = WorkTimeService(db_path)

    opened = service.open_day(telegram_id=1001, now=datetime(2026, 6, 15, 7, 12))
    assert opened.ok
    closed = service.close_open_day(telegram_id=1001, duration_minutes=600)

    assert closed.ok
    assert closed.day is not None
    assert closed.day.start_time == '07:12'
    assert closed.day.end_time == '17:12'
    assert closed.day.total_minutes == 600
    assert closed.day.status == 'closed'


def test_manual_range_requires_no_conflicting_same_day_entry(tmp_path: Path) -> None:
    db_path = tmp_path / 'test.db'
    init_db(db_path)
    service = WorkTimeService(db_path)
    candidate = parse_manual_range_candidate('pracoval som dnes od 5:30 do 17:00', today=date(2026, 6, 15))
    assert candidate is not None

    first = service.add_manual_range(telegram_id=1001, candidate=candidate)
    second = service.add_manual_range(telegram_id=1001, candidate=candidate)

    assert first.ok
    assert not second.ok
    assert second.reason == 'conflict_same_day'


def test_previous_open_day_blocks_new_open_day(tmp_path: Path) -> None:
    db_path = tmp_path / 'test.db'
    init_db(db_path)
    service = WorkTimeService(db_path)

    assert service.open_day(telegram_id=1001, now=datetime(2026, 6, 14, 7, 0)).ok
    blocked = service.open_day(telegram_id=1001, now=datetime(2026, 6, 15, 7, 0))

    assert not blocked.ok
    assert blocked.reason == 'previous_open_day'
    assert blocked.conflict_day is not None
    assert blocked.conflict_day.work_date == '2026-06-14'


def test_duration_only_entry_stores_total_without_start_end(tmp_path: Path) -> None:
    db_path = tmp_path / 'test.db'
    init_db(db_path)
    service = WorkTimeService(db_path)
    candidate = parse_duration_entry_candidate('dnes 9,5 hodin', today=date(2026, 7, 2))
    assert candidate is not None

    result = service.add_duration_entry(telegram_id=1001, candidate=candidate)

    assert result.ok
    assert result.day is not None
    assert result.day.work_date == '2026-07-02'
    assert result.day.start_time is None
    assert result.day.end_time is None
    assert result.day.total_minutes == 570
    assert format_candidate_preview(candidate) == (
        'Datum: 02.07.2026\n'
        'Typ: pocet hodin bez presneho prichodu/odchodu\n'
        'Prichod: -\n'
        'Odchod: -\n'
        'Hodiny: 9,5 hod.'
    )
    assert format_day_summary(result.day) == '02.07.2026: bez presneho prichodu/odchodu (9,5 hod.)'



def test_delete_month_removes_only_current_user_selected_month(tmp_path: Path) -> None:
    db_path = tmp_path / 'test.db'
    init_db(db_path)
    service = WorkTimeService(db_path)

    july_user_a = WorkTimeCandidate(work_date=date(2026, 7, 2), duration_minutes=600, close_mode='manual_duration')
    july_user_b = WorkTimeCandidate(work_date=date(2026, 7, 3), duration_minutes=480, close_mode='manual_duration')
    august_user_a = WorkTimeCandidate(work_date=date(2026, 8, 2), duration_minutes=300, close_mode='manual_duration')
    assert service.add_duration_entry(telegram_id=1001, candidate=july_user_a).ok
    assert service.add_duration_entry(telegram_id=2002, candidate=july_user_b).ok
    assert service.add_duration_entry(telegram_id=1001, candidate=august_user_a).ok

    preview = service.summarize_month(telegram_id=1001, year=2026, month=7)
    assert preview.row_count == 1
    assert preview.total_minutes == 600
    assert service.list_days_for_month(telegram_id=1001, year=2026, month=7)

    deleted = service.delete_month(telegram_id=1001, year=2026, month=7)

    assert deleted.row_count == 1
    assert deleted.total_minutes == 600
    assert service.list_days_for_month(telegram_id=1001, year=2026, month=7) == []
    assert len(service.list_days_for_month(telegram_id=2002, year=2026, month=7)) == 1
    assert len(service.list_days_for_month(telegram_id=1001, year=2026, month=8)) == 1


def test_parse_explicit_month_requires_month_for_delete() -> None:
    assert parse_explicit_month('vymaz dochadzku za jul', today=date(2026, 7, 2)) == (2026, 7)
    assert parse_explicit_month('vymaz dochadzku za 2026-07', today=date(2026, 7, 2)) == (2026, 7)
    assert parse_explicit_month('vymaz dochadzku', today=date(2026, 7, 2)) is None

def test_duration_only_entry_appears_in_report_without_times(tmp_path: Path) -> None:
    db_path = tmp_path / 'test.db'
    init_db(db_path)
    service = WorkTimeService(db_path)
    candidate = WorkTimeCandidate(work_date=date(2026, 7, 2), duration_minutes=600, close_mode='manual_duration')
    assert service.add_duration_entry(telegram_id=1001, candidate=candidate).ok

    report = service.generate_monthly_report(
        telegram_id=1001,
        year=2026,
        month=7,
        output_dir=tmp_path / 'reports',
    )

    workbook = load_workbook(report.report_path)
    sheet = workbook.active
    assert [sheet[f'{column}5'].value for column in 'ABCD'] == ['02.07.2026', None, None, '10 hod.']
    assert sheet['D35'].value == '10 hod.'


def test_report_includes_all_days_sundays_and_total(tmp_path: Path) -> None:
    db_path = tmp_path / 'test.db'
    init_db(db_path)
    service = WorkTimeService(db_path)
    candidate = parse_manual_range_candidate('pracoval som dnes od 08:00 do 17:00', today=date(2026, 7, 1))
    assert candidate is not None
    assert service.add_manual_range(telegram_id=1001, candidate=candidate).ok

    report = service.generate_monthly_report(
        telegram_id=1001,
        year=2026,
        month=7,
        output_dir=tmp_path / 'reports',
    )

    assert report.ok
    assert report.report_path is not None
    assert report.report_path.name == 'dochadzka_2026_07.xlsx'
    workbook = load_workbook(report.report_path)
    sheet = workbook.active
    assert sheet.max_column == 4
    assert sheet.max_row == 35  # title, blank, header, 31 days, total
    assert sheet.sheet_view.showGridLines is False
    assert sheet.print_area == "'Dochadzka'!$A$1:$D$35"
    assert 'A1:D1' in {str(range_) for range_ in sheet.merged_cells.ranges}
    assert sheet['A1'].value == 'Dochádzka — júl 2026'
    assert sheet['A1'].font.bold
    assert sheet['A1'].font.sz == 13
    assert [sheet[f'{column}3'].value for column in 'ABCD'] == ['Dátum', 'Príchod', 'Odchod', 'Hodiny']
    assert all(sheet[f'{column}3'].font.bold for column in 'ABCD')
    assert all(sheet[f'{column}3'].border.bottom.style == 'medium' for column in 'ABCD')
    assert sheet['A4'].value == '01.07.2026'
    assert sheet['B4'].value == '08:00'
    assert sheet['C4'].value == '17:00'
    assert sheet['D4'].value == '9 hod.'
    assert sheet['A35'].value == 'Spolu'
    assert sheet['D35'].value == '9 hod.'
    assert sheet['A35'].font.bold
    assert sheet['D35'].font.bold
    assert all(sheet[f'{column}35'].border.top.style == 'medium' for column in 'ABCD')
    sunday_row = 8  # 05.07.2026
    assert all(sheet[f'{column}{sunday_row}'].fill.fgColor.rgb in {'00C6EFCE', 'C6EFCE'} for column in 'ABCD')
    assert [sheet.column_dimensions[column].width for column in 'ABCD'] == [14, 11, 11, 14]



def test_lunch_break_parser_accepts_supported_examples() -> None:
    assert parse_lunch_break_minutes('30') == 30
    assert parse_lunch_break_minutes('45 minut') == 45
    assert parse_lunch_break_minutes('1 hodina') == 60
    assert parse_lunch_break_minutes('hodina') == 60
    assert parse_lunch_break_minutes('\u043f\u0456\u0432 \u0433\u043e\u0434\u0438\u043d\u0438') == 30
    assert parse_lunch_break_minutes('\u0447\u0430\u0441') == 60
    assert parse_lunch_break_minutes('60 \u0445\u0432\u0438\u043b\u0438\u043d') == 60
    assert parse_lunch_break_minutes('-5') is None
    assert parse_lunch_break_minutes('240 minut') is None


def test_explicit_range_report_subtracts_current_lunch_break(tmp_path: Path) -> None:
    db_path = tmp_path / 'test.db'
    init_db(db_path)
    service = WorkTimeService(db_path)
    service.save_lunch_break_settings(telegram_id=1001, enabled=True, minutes=60)
    candidate = WorkTimeCandidate(work_date=date(2026, 7, 2), start_time=datetime.strptime('09:00', '%H:%M').time(), end_time=datetime.strptime('21:00', '%H:%M').time())
    result = service.add_manual_range(telegram_id=1001, candidate=candidate)
    assert result.ok
    assert result.day is not None
    assert result.day.gross_minutes == 720
    assert result.day.total_minutes == 660

    report = service.generate_monthly_report(telegram_id=1001, year=2026, month=7, output_dir=tmp_path / 'reports')
    sheet = load_workbook(report.report_path).active
    assert sheet['D5'].value == '11 hod.'
    assert sheet['D35'].value == '11 hod.'

    service.save_lunch_break_settings(telegram_id=1001, enabled=True, minutes=30)
    updated_report = service.generate_monthly_report(telegram_id=1001, year=2026, month=7, output_dir=tmp_path / 'reports2')
    updated_sheet = load_workbook(updated_report.report_path).active
    assert updated_sheet['D5'].value == '11,5 hod.'


def test_duration_close_keeps_requested_net_and_extends_end_by_lunch(tmp_path: Path) -> None:
    db_path = tmp_path / 'test.db'
    init_db(db_path)
    service = WorkTimeService(db_path)
    service.save_lunch_break_settings(telegram_id=1001, enabled=True, minutes=60)
    assert service.open_day(telegram_id=1001, now=datetime(2026, 7, 2, 7, 12)).ok

    closed = service.close_open_day(telegram_id=1001, duration_minutes=600)

    assert closed.ok
    assert closed.day is not None
    assert closed.day.end_time == '18:12'
    assert closed.day.gross_minutes == 660
    assert closed.day.total_minutes == 600
    assert closed.day.net_work_minutes_override == 600
    report = service.generate_monthly_report(telegram_id=1001, year=2026, month=7, output_dir=tmp_path / 'reports')
    sheet = load_workbook(report.report_path).active
    assert sheet['D5'].value == '10 hod.'

    service.save_lunch_break_settings(telegram_id=1001, enabled=True, minutes=30)
    updated_report = service.generate_monthly_report(telegram_id=1001, year=2026, month=7, output_dir=tmp_path / 'reports2')
    assert load_workbook(updated_report.report_path).active['D5'].value == '10 hod.'


def test_duration_close_with_lunch_disabled_keeps_old_end_semantics(tmp_path: Path) -> None:
    db_path = tmp_path / 'test.db'
    init_db(db_path)
    service = WorkTimeService(db_path)
    service.save_lunch_break_settings(telegram_id=1001, enabled=False, minutes=0)
    assert service.open_day(telegram_id=1001, now=datetime(2026, 7, 2, 7, 12)).ok

    closed = service.close_open_day(telegram_id=1001, duration_minutes=600)

    assert closed.ok
    assert closed.day is not None
    assert closed.day.end_time == '17:12'
    assert closed.day.total_minutes == 600


def test_report_clamps_break_larger_than_explicit_gross_to_zero(tmp_path: Path) -> None:
    db_path = tmp_path / 'test.db'
    init_db(db_path)
    service = WorkTimeService(db_path)
    service.save_lunch_break_settings(telegram_id=1001, enabled=True, minutes=180)
    candidate = WorkTimeCandidate(work_date=date(2026, 7, 2), start_time=datetime.strptime('09:00', '%H:%M').time(), end_time=datetime.strptime('10:00', '%H:%M').time())
    assert service.add_manual_range(telegram_id=1001, candidate=candidate).ok
    report = service.generate_monthly_report(telegram_id=1001, year=2026, month=7, output_dir=tmp_path / 'reports')
    assert load_workbook(report.report_path).active['D5'].value == '0 hod.'

def test_format_duration_uses_compact_hour_labels() -> None:
    assert _format_duration(540) == '9 hod.'
    assert _format_duration(570) == '9,5 hod.'
    assert _format_duration(6000) == '100 hod.'
    assert _format_duration(6030) == '100,5 hod.'


def test_parsers_separate_duration_from_clock_time() -> None:
    open_day = WorkTimeService.__new__(WorkTimeService)  # type: ignore[misc]
    del open_day
    from bot.services.work_time import WorkTimeDay

    day = WorkTimeDay(1, 1001, '2026-06-15', '07:12', None, None, 'open', 'opened_live', None)
    duration = parse_close_candidate('zatvor den 10 hodin', open_day=day)
    explicit = parse_close_candidate('zatvor dnes o 17:00', open_day=day)

    assert duration is not None
    assert duration.duration_minutes == 600
    assert duration.close_mode == 'close_with_duration'
    assert explicit is not None
    assert explicit.end_time is not None
    assert explicit.end_time.strftime('%H:%M') == '17:00'
    assert explicit.close_mode == 'close_at_time'


def test_report_month_defaults_year_and_month() -> None:
    assert parse_report_month('vytvor vykaz hodin za jun', today=date(2026, 7, 1)) == (2026, 6)
    assert parse_report_month('sformuj dochadzku za jun 2025', today=date(2026, 7, 1)) == (2025, 6)


def test_llm_slot_extractor_normalizes_human_spoken_range(monkeypatch) -> None:
    _WorkTimeSlotOpenAIFake.output = json.dumps(
        {
            'canonical': 'work_time_entry',
            'date': '2026-07-02',
            'start_time': '05:00',
            'end_time': '09:00',
            'duration_minutes': None,
        }
    )
    monkeypatch.setattr('bot.services.work_time.AsyncOpenAI', _WorkTimeSlotOpenAIFake)

    candidate = asyncio.run(
        resolve_work_time_entry_candidate(
            user_input_text="worked today from fifth morning to ninth morning",
            api_key='sk-test',
            model='gpt-4o',
            today=date(2026, 7, 2),
        )
    )

    assert candidate is not None
    assert candidate.work_date == date(2026, 7, 2)
    assert candidate.start_time is not None
    assert candidate.start_time.strftime('%H:%M') == '05:00'
    assert candidate.end_time is not None
    assert candidate.end_time.strftime('%H:%M') == '09:00'
    assert _WorkTimeSlotOpenAIFake.last_payload is not None
    assert _WorkTimeSlotOpenAIFake.last_payload['context_name'] == 'work_time_slot_extraction'


def test_llm_slot_extractor_normalizes_duration_only(monkeypatch) -> None:
    _WorkTimeSlotOpenAIFake.output = json.dumps(
        {
            'canonical': 'work_time_entry',
            'date': '2026-07-02',
            'start_time': None,
            'end_time': None,
            'duration_minutes': 570,
        }
    )
    monkeypatch.setattr('bot.services.work_time.AsyncOpenAI', _WorkTimeSlotOpenAIFake)

    candidate = asyncio.run(
        resolve_work_time_entry_candidate(
            user_input_text="record today nine and a half hours",
            api_key='sk-test',
            model='gpt-4o',
            today=date(2026, 7, 2),
        )
    )

    assert candidate is not None
    assert candidate.work_date == date(2026, 7, 2)
    assert candidate.start_time is None
    assert candidate.end_time is None
    assert candidate.duration_minutes == 570
    assert format_candidate_preview(candidate) == (
        'Datum: 02.07.2026\n'
        'Typ: pocet hodin bez presneho prichodu/odchodu\n'
        'Prichod: -\n'
        'Odchod: -\n'
        'Hodiny: 9,5 hod.'
    )



def test_manual_range_relative_date_yesterday_variants() -> None:
    examples = [
        'včera od 6:00 do 16:30',
        'vcera od 6:00 do 16:30',
        'вчора з 6:00 до 18:00',
        'учора з 6:00 до 18:00',
        'вчера с 6:00 до 18:00',
    ]
    for text in examples:
        candidate = parse_manual_range_candidate(text, today=date(2026, 7, 2))
        assert candidate is not None, text
        assert candidate.work_date == date(2026, 7, 1)


def test_lunch_math_regressions_stay_net_gross_safe(tmp_path: Path) -> None:
    db_path = tmp_path / 'test.db'
    init_db(db_path)
    service = WorkTimeService(db_path)
    service.save_lunch_break_settings(telegram_id=1001, enabled=True, minutes=60)

    explicit = parse_manual_range_candidate('dnes od 9:00 do 21:00', today=date(2026, 7, 2))
    assert explicit is not None
    saved = service.add_manual_range(telegram_id=1001, candidate=explicit)
    assert saved.ok
    assert saved.day is not None
    assert service.report_net_minutes(saved.day, settings=service.get_lunch_break_settings(telegram_id=1001)) == 660

    deleted = service.delete_month(telegram_id=1001, year=2026, month=7)
    assert deleted.row_count == 1
    opened = service.open_day(telegram_id=1001, now=datetime(2026, 7, 2, 7, 12))
    assert opened.ok
    closed = service.close_open_day(telegram_id=1001, duration_minutes=600)
    assert closed.ok
    assert closed.day is not None
    assert closed.day.end_time == '18:12'
    assert service.report_net_minutes(closed.day, settings=service.get_lunch_break_settings(telegram_id=1001)) == 600
