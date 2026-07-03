from __future__ import annotations

from calendar import monthrange
from dataclasses import dataclass
from datetime import UTC, date, datetime, time, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError
import json
import logging
import os
import re
import sqlite3
import unicodedata

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openai import AsyncOpenAI

from bot.services.db import managed_connection


STATUS_OPEN = 'open'
STATUS_CLOSED = 'closed'
STATUS_SKIPPED = 'skipped'
STATUS_DAY_OFF = 'day_off'
STATUS_NON_WORKING = 'non_working'

SOURCE_OPENED_LIVE = 'opened_live'
SOURCE_CLOSED_LIVE = 'closed_live'
SOURCE_MANUAL_RANGE = 'manual_range'
SOURCE_MANUAL_DURATION = 'manual_duration'
SOURCE_GENERATED_CALENDAR = 'generated_calendar'
DEFAULT_WORK_TIME_TIMEZONE = 'Europe/Bratislava'
_WORK_TIME_TIMEZONE_ENV = 'OFFICEFLOW_TIMEZONE'
_LOGGER = logging.getLogger(__name__)


MONTH_NAMES_SK = {
    1: 'január',
    2: 'február',
    3: 'marec',
    4: 'apríl',
    5: 'máj',
    6: 'jún',
    7: 'júl',
    8: 'august',
    9: 'september',
    10: 'október',
    11: 'november',
    12: 'december',
}


@dataclass(frozen=True)
class WorkTimeDay:
    id: int
    telegram_id: int
    work_date: str
    start_time: str | None
    end_time: str | None
    total_minutes: int | None
    status: str
    source: str
    note: str | None
    gross_minutes: int | None = None
    lunch_break_minutes_snapshot: int | None = None
    net_work_minutes_override: int | None = None
    close_input_mode: str | None = None


@dataclass(frozen=True)
class WorkTimeLunchBreakSettings:
    telegram_id: int
    configured: bool
    enabled: bool
    minutes: int


@dataclass(frozen=True)
class WorkTimeOperationResult:
    ok: bool
    day: WorkTimeDay | None = None
    reason: str = ''
    conflict_day: WorkTimeDay | None = None
    report_path: Path | None = None


@dataclass(frozen=True)
class WorkTimeMonthSummary:
    year: int
    month: int
    row_count: int
    total_minutes: int


@dataclass(frozen=True)
class WorkTimeCandidate:
    work_date: date
    start_time: time | None = None
    end_time: time | None = None
    duration_minutes: int | None = None
    close_mode: str = 'unknown'
    needs_confirmation: bool = True

    @property
    def calculated_end_time(self) -> time | None:
        if self.end_time is not None:
            return self.end_time
        if self.start_time is not None and self.duration_minutes is not None:
            base = datetime.combine(self.work_date, self.start_time)
            return (base + timedelta(minutes=self.duration_minutes)).time().replace(second=0, microsecond=0)
        return None


class WorkTimeService:
    def __init__(self, db_path: Path) -> None:
        self.db_path = db_path

    def open_day(
        self,
        *,
        telegram_id: int,
        now: datetime | None = None,
        source_message_id: int | None = None,
    ) -> WorkTimeOperationResult:
        current = _local_now(now)
        work_date = current.date().isoformat()
        start_value = _format_time(current.time())
        with managed_connection(self.db_path) as connection:
            open_day = self.get_open_day(telegram_id=telegram_id, connection=connection)
            if open_day is not None:
                if open_day.work_date == work_date:
                    return WorkTimeOperationResult(ok=True, day=open_day, reason='already_open')
                return WorkTimeOperationResult(ok=False, reason='previous_open_day', conflict_day=open_day)

            existing = self.get_day(telegram_id=telegram_id, work_date=work_date, connection=connection)
            if existing is not None:
                return WorkTimeOperationResult(ok=False, reason=f'already_{existing.status}', conflict_day=existing)

            cursor = connection.execute(
                """
                INSERT INTO work_time_days
                    (telegram_id, work_date, start_time, end_time, total_minutes, status, source, note, close_input_mode)
                VALUES (?, ?, ?, NULL, NULL, ?, ?, NULL, NULL)
                """,
                (telegram_id, work_date, start_value, STATUS_OPEN, SOURCE_OPENED_LIVE),
            )
            day_id = int(cursor.lastrowid)
            self._record_event(
                connection,
                day_id=day_id,
                event_type='open',
                old_value=None,
                new_value={'start_time': start_value, 'work_date': work_date},
                source_message_id=source_message_id,
                telegram_id=telegram_id,
            )
            connection.commit()
            return WorkTimeOperationResult(ok=True, day=self.get_day_by_id(day_id, connection=connection))

    def close_open_day(
        self,
        *,
        telegram_id: int,
        end_datetime: datetime | None = None,
        duration_minutes: int | None = None,
        source: str = SOURCE_CLOSED_LIVE,
        source_message_id: int | None = None,
    ) -> WorkTimeOperationResult:
        current = _local_now(end_datetime)
        with managed_connection(self.db_path) as connection:
            day = self.get_open_day(telegram_id=telegram_id, connection=connection)
            if day is None:
                return WorkTimeOperationResult(ok=False, reason='no_open_day')
            if day.start_time is None:
                return WorkTimeOperationResult(ok=False, reason='missing_start_time', conflict_day=day)

            settings = self.get_lunch_break_settings(telegram_id=telegram_id, connection=connection)
            effective_break = _effective_lunch_break_minutes(settings)
            start_dt = datetime.combine(date.fromisoformat(day.work_date), _parse_time_value(day.start_time))
            if duration_minutes is not None:
                if duration_minutes <= 0:
                    return WorkTimeOperationResult(ok=False, reason='invalid_duration', conflict_day=day)
                end_dt = start_dt + timedelta(minutes=duration_minutes + effective_break)
                gross_minutes = duration_minutes + effective_break
                total_minutes = duration_minutes
                net_override = duration_minutes
                lunch_snapshot = effective_break
                close_mode = 'duration'
                close_source = SOURCE_MANUAL_DURATION
            else:
                end_dt = datetime.combine(date.fromisoformat(day.work_date), current.time().replace(second=0, microsecond=0))
                if end_dt < start_dt:
                    return WorkTimeOperationResult(ok=False, reason='end_before_start', conflict_day=day)
                gross_minutes = int((end_dt - start_dt).total_seconds() // 60)
                total_minutes = max(0, gross_minutes - effective_break)
                net_override = None
                lunch_snapshot = None
                close_mode = 'explicit_end'
                close_source = source
            if end_dt < start_dt:
                return WorkTimeOperationResult(ok=False, reason='end_before_start', conflict_day=day)
            end_value = _format_time(end_dt.time())
            connection.execute(
                """
                UPDATE work_time_days
                SET end_time = ?, total_minutes = ?, status = ?, source = ?, gross_minutes = ?,
                    lunch_break_minutes_snapshot = ?, net_work_minutes_override = ?, close_input_mode = ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ? AND telegram_id = ?
                """,
                (
                    end_value,
                    total_minutes,
                    STATUS_CLOSED,
                    close_source,
                    gross_minutes,
                    lunch_snapshot,
                    net_override,
                    close_mode,
                    day.id,
                    telegram_id,
                ),
            )
            self._record_event(
                connection,
                day_id=day.id,
                event_type='close',
                old_value=day.__dict__,
                new_value={
                    'end_time': end_value,
                    'gross_minutes': gross_minutes,
                    'total_minutes': total_minutes,
                    'lunch_break_minutes_snapshot': lunch_snapshot,
                    'net_work_minutes_override': net_override,
                    'close_input_mode': close_mode,
                    'source': close_source,
                },
                source_message_id=source_message_id,
                telegram_id=telegram_id,
            )
            connection.commit()
            return WorkTimeOperationResult(ok=True, day=self.get_day_by_id(day.id, connection=connection))

    def add_manual_range(
        self,
        *,
        telegram_id: int,
        candidate: WorkTimeCandidate,
        source_message_id: int | None = None,
    ) -> WorkTimeOperationResult:
        if candidate.start_time is None or candidate.calculated_end_time is None:
            return WorkTimeOperationResult(ok=False, reason='missing_range')
        start_dt = datetime.combine(candidate.work_date, candidate.start_time)
        end_dt = datetime.combine(candidate.work_date, candidate.calculated_end_time)
        if end_dt < start_dt:
            return WorkTimeOperationResult(ok=False, reason='end_before_start')
        gross_minutes = int((end_dt - start_dt).total_seconds() // 60)
        with managed_connection(self.db_path) as connection:
            settings = self.get_lunch_break_settings(telegram_id=telegram_id, connection=connection)
            total_minutes = max(0, gross_minutes - _effective_lunch_break_minutes(settings))
            existing = self.get_day(
                telegram_id=telegram_id,
                work_date=candidate.work_date.isoformat(),
                connection=connection,
            )
            if existing is not None and existing.status in {STATUS_OPEN, STATUS_CLOSED}:
                return WorkTimeOperationResult(ok=False, reason='conflict_same_day', conflict_day=existing)
            if existing is not None:
                connection.execute('DELETE FROM work_time_days WHERE id = ? AND telegram_id = ?', (existing.id, telegram_id))
            cursor = connection.execute(
                """
                INSERT INTO work_time_days
                    (telegram_id, work_date, start_time, end_time, total_minutes, status, source, note, gross_minutes, close_input_mode)
                VALUES (?, ?, ?, ?, ?, ?, ?, NULL, ?, ?)
                """,
                (
                    telegram_id,
                    candidate.work_date.isoformat(),
                    _format_time(candidate.start_time),
                    _format_time(candidate.calculated_end_time),
                    total_minutes,
                    STATUS_CLOSED,
                    SOURCE_MANUAL_RANGE,
                    gross_minutes,
                    'explicit_range',
                ),
            )
            day_id = int(cursor.lastrowid)
            self._record_event(
                connection,
                day_id=day_id,
                event_type='open',
                old_value=None,
                new_value={'manual_range': True, 'gross_minutes': gross_minutes, 'total_minutes': total_minutes, 'close_input_mode': 'explicit_range'},
                source_message_id=source_message_id,
                telegram_id=telegram_id,
            )
            connection.commit()
            return WorkTimeOperationResult(ok=True, day=self.get_day_by_id(day_id, connection=connection))

    def add_duration_entry(
        self,
        *,
        telegram_id: int,
        candidate: WorkTimeCandidate,
        source_message_id: int | None = None,
    ) -> WorkTimeOperationResult:
        if candidate.duration_minutes is None or candidate.duration_minutes <= 0:
            return WorkTimeOperationResult(ok=False, reason='invalid_duration')
        with managed_connection(self.db_path) as connection:
            existing = self.get_day(
                telegram_id=telegram_id,
                work_date=candidate.work_date.isoformat(),
                connection=connection,
            )
            if existing is not None and existing.status in {STATUS_OPEN, STATUS_CLOSED}:
                return WorkTimeOperationResult(ok=False, reason='conflict_same_day', conflict_day=existing)
            if existing is not None:
                connection.execute('DELETE FROM work_time_days WHERE id = ? AND telegram_id = ?', (existing.id, telegram_id))
            lunch_snapshot = _effective_lunch_break_minutes(self.get_lunch_break_settings(telegram_id=telegram_id, connection=connection))
            cursor = connection.execute(
                """
                INSERT INTO work_time_days
                    (telegram_id, work_date, start_time, end_time, total_minutes, status, source, note,
                     lunch_break_minutes_snapshot, net_work_minutes_override, close_input_mode)
                VALUES (?, ?, NULL, NULL, ?, ?, ?, NULL, ?, ?, ?)
                """,
                (
                    telegram_id,
                    candidate.work_date.isoformat(),
                    candidate.duration_minutes,
                    STATUS_CLOSED,
                    SOURCE_MANUAL_DURATION,
                    lunch_snapshot,
                    candidate.duration_minutes,
                    'duration',
                ),
            )
            day_id = int(cursor.lastrowid)
            self._record_event(
                connection,
                day_id=day_id,
                event_type='duration_entry',
                old_value=None,
                new_value={
                    'duration_only': True,
                    'total_minutes': candidate.duration_minutes,
                    'lunch_break_minutes_snapshot': lunch_snapshot,
                    'net_work_minutes_override': candidate.duration_minutes,
                    'close_input_mode': 'duration',
                },
                source_message_id=source_message_id,
                telegram_id=telegram_id,
            )
            connection.commit()
            return WorkTimeOperationResult(ok=True, day=self.get_day_by_id(day_id, connection=connection))

    def skip_open_day(
        self,
        *,
        telegram_id: int,
        note: str = 'Preskocene pouzivatelom pri otvorenom dni.',
        source_message_id: int | None = None,
    ) -> WorkTimeOperationResult:
        with managed_connection(self.db_path) as connection:
            day = self.get_open_day(telegram_id=telegram_id, connection=connection)
            if day is None:
                return WorkTimeOperationResult(ok=False, reason='no_open_day')
            connection.execute(
                """
                UPDATE work_time_days
                SET status = ?, end_time = NULL, total_minutes = NULL, note = ?, updated_at = CURRENT_TIMESTAMP
                WHERE id = ? AND telegram_id = ?
                """,
                (STATUS_SKIPPED, note, day.id, telegram_id),
            )
            self._record_event(
                connection,
                day_id=day.id,
                event_type='skip',
                old_value=day.__dict__,
                new_value={'status': STATUS_SKIPPED, 'note': note},
                source_message_id=source_message_id,
                telegram_id=telegram_id,
            )
            connection.commit()
            return WorkTimeOperationResult(ok=True, day=self.get_day_by_id(day.id, connection=connection))

    def generate_monthly_report(
        self,
        *,
        telegram_id: int,
        year: int,
        month: int,
        output_dir: Path,
    ) -> WorkTimeOperationResult:
        rows = self.list_days_for_month(telegram_id=telegram_id, year=year, month=month)
        settings = self.get_lunch_break_settings(telegram_id=telegram_id)
        by_date = {row.work_date: row for row in rows}
        output_dir.mkdir(parents=True, exist_ok=True)
        report_path = output_dir / f'dochadzka_{year:04d}_{month:02d}.xlsx'
        title = f"Dochádzka — {MONTH_NAMES_SK[month]} {year}"

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Dochadzka'
        sheet['A1'] = title
        sheet['A1'].font = Font(bold=True, size=13)
        sheet.merge_cells('A1:D1')
        sheet.append([])
        sheet.append(['Dátum', 'Príchod', 'Odchod', 'Hodiny (h:mm)'])
        thin_side = Side(style='thin', color='D9E2EC')
        strong_side = Side(style='medium', color='7A8794')
        table_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        header_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=strong_side)
        total_border = Border(left=thin_side, right=thin_side, top=strong_side, bottom=thin_side)
        for cell in sheet[3]:
            cell.font = Font(bold=True)
            cell.border = header_border
            cell.alignment = Alignment(horizontal='center')

        total_minutes = 0
        _, day_count = monthrange(year, month)
        sunday_fill = PatternFill(fill_type='solid', fgColor='C6EFCE')
        for day_number in range(1, day_count + 1):
            current_date = date(year, month, day_number)
            record = by_date.get(current_date.isoformat())
            if record and record.status == STATUS_CLOSED and record.total_minutes is not None:
                report_minutes = self.report_net_minutes(record, settings=settings)
                duration_value = _excel_duration_value(report_minutes)
                total_minutes += report_minutes
                row_values = [current_date.strftime('%d.%m.%Y'), record.start_time or '', record.end_time or '', duration_value]
            elif record and record.status == STATUS_SKIPPED:
                row_values = [current_date.strftime('%d.%m.%Y'), '', '', 'preskocene']
            else:
                row_values = [current_date.strftime('%d.%m.%Y'), '', '', '']
            sheet.append(row_values)
            for cell in sheet[sheet.max_row]:
                cell.border = table_border
            duration_cell = sheet.cell(row=sheet.max_row, column=4)
            if isinstance(duration_cell.value, (int, float)):
                duration_cell.number_format = '[h]:mm'
            if current_date.weekday() == 6:
                for cell in sheet[sheet.max_row]:
                    cell.fill = sunday_fill

        sheet.append(['Spolu', '', '', _excel_duration_value(total_minutes)])
        sheet.cell(row=sheet.max_row, column=4).number_format = '[h]:mm'
        for cell in sheet[sheet.max_row]:
            cell.font = Font(bold=True)
            cell.border = total_border
        sheet.column_dimensions['A'].width = 14
        sheet.column_dimensions['B'].width = 11
        sheet.column_dimensions['C'].width = 11
        sheet.column_dimensions['D'].width = 14
        sheet.freeze_panes = 'A4'
        sheet.sheet_view.showGridLines = False
        sheet.print_area = f'A1:D{sheet.max_row}'
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        workbook.save(report_path)

        with managed_connection(self.db_path) as connection:
            self._record_event(
                connection,
                day_id=None,
                event_type='report_generated',
                old_value=None,
                new_value={'year': year, 'month': month, 'path': str(report_path)},
                source_message_id=None,
                telegram_id=telegram_id,
            )
            connection.commit()
        return WorkTimeOperationResult(ok=True, report_path=report_path)

    def summarize_month(self, *, telegram_id: int, year: int, month: int) -> WorkTimeMonthSummary:
        rows = self.list_days_for_month(telegram_id=telegram_id, year=year, month=month)
        settings = self.get_lunch_break_settings(telegram_id=telegram_id)
        return WorkTimeMonthSummary(
            year=year,
            month=month,
            row_count=len(rows),
            total_minutes=sum(self.report_net_minutes(row, settings=settings) for row in rows),
        )

    def delete_month(
        self,
        *,
        telegram_id: int,
        year: int,
        month: int,
        source_message_id: int | None = None,
    ) -> WorkTimeMonthSummary:
        start = date(year, month, 1).isoformat()
        end = date(year, month, monthrange(year, month)[1]).isoformat()
        with managed_connection(self.db_path) as connection:
            rows = self._list_days_for_month_with_connection(
                connection=connection,
                telegram_id=telegram_id,
                year=year,
                month=month,
            )
            day_ids = [row.id for row in rows]
            settings = self.get_lunch_break_settings(telegram_id=telegram_id, connection=connection)
            summary = WorkTimeMonthSummary(
                year=year,
                month=month,
                row_count=len(day_ids),
                total_minutes=sum(self.report_net_minutes(row, settings=settings) for row in rows),
            )
            if day_ids:
                placeholders = ','.join('?' for _ in day_ids)
                connection.execute(
                    f'DELETE FROM work_time_events WHERE telegram_id = ? AND work_time_day_id IN ({placeholders})',
                    (telegram_id, *day_ids),
                )
                connection.execute(
                    """
                    DELETE FROM work_time_days
                    WHERE telegram_id = ? AND work_date BETWEEN ? AND ?
                    """,
                    (telegram_id, start, end),
                )
                self._record_event(
                    connection,
                    day_id=None,
                    event_type='delete_month',
                    old_value=None,
                    new_value={
                        'year': year,
                        'month': month,
                        'row_count': summary.row_count,
                        'total_minutes': summary.total_minutes,
                    },
                    source_message_id=source_message_id,
                    telegram_id=telegram_id,
                )
            connection.commit()
            return summary

    def get_open_day(
        self,
        *,
        telegram_id: int,
        connection: sqlite3.Connection | None = None,
    ) -> WorkTimeDay | None:
        def _query(conn: sqlite3.Connection) -> WorkTimeDay | None:
            row = conn.execute(
                """
                SELECT id, telegram_id, work_date, start_time, end_time, total_minutes, status, source, note, gross_minutes, lunch_break_minutes_snapshot, net_work_minutes_override, close_input_mode
                FROM work_time_days
                WHERE telegram_id = ? AND status = ?
                ORDER BY work_date ASC, id ASC
                LIMIT 1
                """,
                (telegram_id, STATUS_OPEN),
            ).fetchone()
            return _row_to_day(row)

        if connection is not None:
            return _query(connection)
        with managed_connection(self.db_path) as conn:
            return _query(conn)

    def get_day(
        self,
        *,
        telegram_id: int,
        work_date: str,
        connection: sqlite3.Connection | None = None,
    ) -> WorkTimeDay | None:
        def _query(conn: sqlite3.Connection) -> WorkTimeDay | None:
            row = conn.execute(
                """
                SELECT id, telegram_id, work_date, start_time, end_time, total_minutes, status, source, note, gross_minutes, lunch_break_minutes_snapshot, net_work_minutes_override, close_input_mode
                FROM work_time_days
                WHERE telegram_id = ? AND work_date = ?
                ORDER BY id DESC
                LIMIT 1
                """,
                (telegram_id, work_date),
            ).fetchone()
            return _row_to_day(row)

        if connection is not None:
            return _query(connection)
        with managed_connection(self.db_path) as conn:
            return _query(conn)

    def get_day_by_id(
        self,
        day_id: int,
        *,
        connection: sqlite3.Connection | None = None,
    ) -> WorkTimeDay | None:
        def _query(conn: sqlite3.Connection) -> WorkTimeDay | None:
            row = conn.execute(
                """
                SELECT id, telegram_id, work_date, start_time, end_time, total_minutes, status, source, note, gross_minutes, lunch_break_minutes_snapshot, net_work_minutes_override, close_input_mode
                FROM work_time_days
                WHERE id = ?
                """,
                (day_id,),
            ).fetchone()
            return _row_to_day(row)

        if connection is not None:
            return _query(connection)
        with managed_connection(self.db_path) as conn:
            return _query(conn)

    def list_days_for_month(self, *, telegram_id: int, year: int, month: int) -> list[WorkTimeDay]:
        with managed_connection(self.db_path) as connection:
            return self._list_days_for_month_with_connection(
                connection=connection,
                telegram_id=telegram_id,
                year=year,
                month=month,
            )

    def get_lunch_break_settings(
        self,
        *,
        telegram_id: int,
        connection: sqlite3.Connection | None = None,
    ) -> WorkTimeLunchBreakSettings:
        def _query(conn: sqlite3.Connection) -> WorkTimeLunchBreakSettings:
            row = conn.execute(
                """
                SELECT telegram_id, lunch_break_configured, lunch_break_enabled, lunch_break_minutes
                FROM work_time_settings
                WHERE telegram_id = ?
                """,
                (telegram_id,),
            ).fetchone()
            if row is None:
                return WorkTimeLunchBreakSettings(telegram_id=telegram_id, configured=False, enabled=False, minutes=0)
            return WorkTimeLunchBreakSettings(
                telegram_id=int(row[0]),
                configured=bool(row[1]),
                enabled=bool(row[2]),
                minutes=max(0, int(row[3] or 0)),
            )

        if connection is not None:
            return _query(connection)
        with managed_connection(self.db_path) as conn:
            return _query(conn)

    def save_lunch_break_settings(
        self,
        *,
        telegram_id: int,
        enabled: bool,
        minutes: int,
    ) -> WorkTimeLunchBreakSettings:
        safe_minutes = _validate_lunch_break_minutes(minutes)
        enabled_value = bool(enabled and safe_minutes > 0)
        stored_minutes = safe_minutes if enabled_value else 0
        with managed_connection(self.db_path) as connection:
            connection.execute(
                """
                INSERT INTO work_time_settings
                    (telegram_id, lunch_break_configured, lunch_break_enabled, lunch_break_minutes, updated_at)
                VALUES (?, 1, ?, ?, CURRENT_TIMESTAMP)
                ON CONFLICT(telegram_id) DO UPDATE SET
                    lunch_break_configured = 1,
                    lunch_break_enabled = excluded.lunch_break_enabled,
                    lunch_break_minutes = excluded.lunch_break_minutes,
                    updated_at = CURRENT_TIMESTAMP
                """,
                (telegram_id, 1 if enabled_value else 0, stored_minutes),
            )
            connection.commit()
            return self.get_lunch_break_settings(telegram_id=telegram_id, connection=connection)

    def report_net_minutes(
        self,
        day: WorkTimeDay,
        *,
        settings: WorkTimeLunchBreakSettings | None = None,
    ) -> int:
        if day.net_work_minutes_override is not None:
            return max(0, day.net_work_minutes_override)
        if day.start_time and day.end_time:
            gross = day.gross_minutes
            if gross is None:
                gross = _minutes_between(day.start_time, day.end_time)
            if gross is None:
                return max(0, day.total_minutes or 0)
            return max(0, gross - _effective_lunch_break_minutes(settings))
        return max(0, day.total_minutes or 0)

    def _list_days_for_month_with_connection(
        self,
        *,
        connection: sqlite3.Connection,
        telegram_id: int,
        year: int,
        month: int,
    ) -> list[WorkTimeDay]:
        start = date(year, month, 1).isoformat()
        end = date(year, month, monthrange(year, month)[1]).isoformat()
        rows = connection.execute(
            """
            SELECT id, telegram_id, work_date, start_time, end_time, total_minutes, status, source, note, gross_minutes, lunch_break_minutes_snapshot, net_work_minutes_override, close_input_mode
            FROM work_time_days
            WHERE telegram_id = ? AND work_date BETWEEN ? AND ?
            ORDER BY work_date ASC, id ASC
            """,
            (telegram_id, start, end),
        ).fetchall()
        days = []
        for row in rows:
            day = _row_to_day(row)
            if day is not None:
                days.append(day)
        return days

    def _record_event(
        self,
        connection: sqlite3.Connection,
        *,
        day_id: int | None,
        event_type: str,
        old_value: dict | None,
        new_value: dict | None,
        source_message_id: int | None,
        telegram_id: int | None = None,
    ) -> None:
        connection.execute(
            """
            INSERT INTO work_time_events
                (work_time_day_id, telegram_id, event_type, old_value, new_value, source_message_id)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (
                day_id,
                telegram_id,
                event_type,
                json.dumps(old_value, ensure_ascii=False) if old_value is not None else None,
                json.dumps(new_value, ensure_ascii=False) if new_value is not None else None,
                source_message_id,
            ),
        )



def parse_lunch_break_minutes(text: str) -> int | None:
    normalized = _normalize(text)
    if not normalized:
        return None
    if any(term in normalized for term in ('pol hodiny', 'polhodina', 'polhodiny', 'half hour', 'piv godini', 'piv hodyny', 'polcasa', 'pol casa')):
        return 30
    if any(term in normalized for term in ('hodina', 'jedna hodina', '1 hodina', 'godina', 'chas', 'cas')):
        return 60
    if any(term in normalized for term in ('година', 'одна година', 'годину', 'час')):
        return 60
    if any(term in normalized for term in ('пів години', 'півгодини', 'полчаса', 'пол часа')):
        return 30
    hour_match = re.search(r'\b(\d{1,2})\s*(?:hodin|hodiny|hodina|hod|h|godin|casov|chas|годин|години|час|часа|часов)\b', normalized)
    if hour_match:
        value = int(hour_match.group(1)) * 60
        return value if 0 <= value <= 180 else None
    minute_match = re.search(r'\b(\d{1,3})\s*(?:minut|minuta|minuty|min|m|хвилин|хв|минут|минута|минуты)\b', normalized)
    if minute_match:
        value = int(minute_match.group(1))
        return value if 0 <= value <= 180 else None
    plain_match = re.fullmatch(r'\s*(\d{1,3})\s*', normalized)
    if plain_match:
        value = int(plain_match.group(1))
        return value if 0 <= value <= 180 else None
    return None


def is_lunch_break_disable_request(text: str) -> bool:
    normalized = _normalize(text)
    if not normalized:
        return False
    disable_terms = ('zrus', 'vypni', 'bez', 'neodpocitavaj', 'ne odnimati', 'ne vidnimati', 'otkluci', 'disable')
    lunch_terms = ('obed', 'obednu', 'obednej', 'pauz', 'prestav', 'lunch', 'break', 'обід', 'обед', 'пауз', 'перерв')
    return any(term in normalized for term in disable_terms) and any(term in normalized for term in lunch_terms)
def parse_manual_range_candidate(text: str, *, today: date | None = None) -> WorkTimeCandidate | None:
    normalized = _normalize(text)
    work_date, parse_text = _resolve_work_date_and_strip_explicit_date(normalized, today=today)
    matches = re.findall(r'\b(\d{1,2})(?::|\.)(\d{2})\b', parse_text)
    if len(matches) < 2:
        return None
    start = _parse_match_time(matches[0])
    end = _parse_match_time(matches[1])
    if start is None or end is None:
        return None
    return WorkTimeCandidate(work_date=work_date, start_time=start, end_time=end, close_mode='manual_range')


def parse_duration_entry_candidate(text: str, *, today: date | None = None) -> WorkTimeCandidate | None:
    normalized = _normalize(text)
    work_date, parse_text = _resolve_work_date_and_strip_explicit_date(normalized, today=today)
    duration_minutes = _parse_duration_minutes(parse_text)
    if duration_minutes is None:
        return None
    return WorkTimeCandidate(
        work_date=work_date,
        duration_minutes=duration_minutes,
        close_mode='manual_duration',
    )

def parse_close_candidate(text: str, *, open_day: WorkTimeDay, today: date | None = None) -> WorkTimeCandidate | None:
    normalized = _normalize(text)
    work_date = date.fromisoformat(open_day.work_date)
    duration_match = re.search(
        r'\b(\d{1,2})(?:[,.](\d{1,2}))?\s*(?:hodin|hodiny|hodina|hod|h|godin|chas|casov|С‡Р°СЃ|С‡Р°СЃРѕРІ|РіРѕРґРёРЅ)\b',
        normalized,
    )
    if duration_match:
        hours = int(duration_match.group(1))
        fraction = duration_match.group(2)
        minutes = 0
        if fraction:
            minutes = int(round(float(f'0.{fraction}') * 60))
        return WorkTimeCandidate(work_date=work_date, duration_minutes=hours * 60 + minutes, close_mode='close_with_duration')

    time_match = re.search(r'(?:\bo\b|\bv\b|\bat\b|РІ|Рѕ)\s*(\d{1,2})(?::(\d{2}))\b', normalized)
    if time_match:
        end_time = _parse_match_time((time_match.group(1), time_match.group(2)))
        if end_time is not None:
            return WorkTimeCandidate(work_date=work_date, end_time=end_time, close_mode='close_at_time')

    if any(term in normalized for term in ('teraz', 'now', 'zaraz', 'sejcas', 'зараз', 'сейчас', 'Р·Р°СЂР°Р·', 'СЃРµР№С‡Р°СЃ')):
        return WorkTimeCandidate(work_date=_resolve_relative_date(normalized, today=today), close_mode='close_now', needs_confirmation=False)
    return None


def parse_report_month(text: str, *, today: date | None = None) -> tuple[int, int]:
    current = today or work_time_local_date()
    normalized = _normalize(text)
    explicit_year = re.findall(r'\b((?:19|20)\d{2})\b', normalized)
    year = int(explicit_year[-1]) if explicit_year else current.year
    for month, names in _MONTH_ALIASES.items():
        if any(name in normalized for name in names):
            return year, month
    if 'minuly mesiac' in normalized or 'last month' in normalized:
        previous = (current.replace(day=1) - timedelta(days=1))
        return previous.year, previous.month
    return year, current.month


def parse_explicit_month(text: str, *, today: date | None = None) -> tuple[int, int] | None:
    current = today or work_time_local_date()
    normalized = _normalize(text)
    explicit_year = re.findall(r'\b((?:19|20)\d{2})\b', normalized)
    year = int(explicit_year[-1]) if explicit_year else current.year
    numeric_year_month = re.search(r'\b((?:19|20)\d{2})[-/.](0?[1-9]|1[0-2])\b', normalized)
    if numeric_year_month:
        return int(numeric_year_month.group(1)), int(numeric_year_month.group(2))
    numeric_month_year = re.search(r'\b(0?[1-9]|1[0-2])[-/.]((?:19|20)\d{2})\b', normalized)
    if numeric_month_year:
        return int(numeric_month_year.group(2)), int(numeric_month_year.group(1))
    for month, names in _MONTH_ALIASES.items():
        if any(name in normalized for name in names):
            return year, month
    if 'minuly mesiac' in normalized or 'last month' in normalized:
        previous = current.replace(day=1) - timedelta(days=1)
        return previous.year, previous.month
    return None


async def resolve_work_time_entry_candidate(
    *,
    user_input_text: str,
    api_key: str | None,
    model: str,
    today: date | None = None,
    open_day: WorkTimeDay | None = None,
) -> WorkTimeCandidate | None:
    cleaned = user_input_text.strip()
    if not cleaned or not api_key or not api_key.startswith('sk-'):
        return None
    current = today or work_time_local_date()
    try:
        client = AsyncOpenAI(api_key=api_key)
        response = await client.chat.completions.create(
            model=model,
            temperature=0,
            response_format={'type': 'json_object'},
            messages=[
                {
                    'role': 'system',
                    'content': (
                        'You are a bounded slot extractor for OfficeFlow work-time tracking. '
                        'Return strict JSON only. Supported input languages are Slovak, Ukrainian, Russian, English, and mixed STT-noisy text. '
                        'Extract only candidate work-time slots; never claim anything was saved. '
                        'Allowed shape: {"canonical":"work_time_entry","mode":"manual_range|manual_duration|close_at_time|close_with_duration|close_now|unknown","date":"YYYY-MM-DD|null","start_time":"HH:MM|null","end_time":"HH:MM|null","duration_minutes":<integer|null>} '
                        'or {"canonical":"unknown"}. '
                        'Use mode=close_now only when the user clearly says closing now/teraz/now/zaraz/sejcas or equivalent. '
                        'Use mode=close_at_time only when closing an already open day by exact end time. Use mode=close_with_duration only when closing an open day by worked duration. '
                        'Use mode=manual_range when the user gives a standalone work interval with start and end time. Use mode=manual_duration only when the user gives a standalone total duration. '
                        'If the user says today/dnes/ukrainian today, use today_iso. If the user gives an explicit date such as 1.07, 1.07.2026, or 1 na 7, normalize it to YYYY-MM-DD using today_iso year when year is missing. If a work day is already open, use open_day_date. '
                        'Verbal ranges such as "z piatej do deviatej", "from fifth morning to ninth morning", "z siestej rano do piatej vecer", or Cyrillic equivalents must be normalized to start_time and end_time in 24-hour HH:MM when clear. '
                        'For duration-only requests like "zapis dnes 10 hodin" or "close today nine and a half hours", set start_time and end_time to null and duration_minutes to the total. Do not convert a range into duration-only just because one number looks like hours. '
                        'If the input is a question about whether the bot can do work time tracking, return unknown rather than an executable work-time entry. Use unknown when date/time/duration is ambiguous.'
                    ),
                },
                {
                    'role': 'user',
                    'content': json.dumps(
                        {
                            'context_name': 'work_time_slot_extraction',
                            'today_iso': current.isoformat(),
                            'open_day_date': open_day.work_date if open_day else None,
                            'open_day_start_time': open_day.start_time if open_day else None,
                            'user_input_text': cleaned,
                            'expected_output': {
                                'canonical': 'work_time_entry or unknown',
                                'mode': 'manual_range, manual_duration, close_at_time, close_with_duration, close_now, or unknown',
                                'date': 'YYYY-MM-DD or null',
                                'start_time': 'HH:MM or null',
                                'end_time': 'HH:MM or null',
                                'duration_minutes': 'positive integer or null',
                            },
                        },
                        ensure_ascii=False,
                    ),
                },
            ],
        )
        parsed = json.loads(response.choices[0].message.content or '{}')
    except Exception:
        return None

    if str(parsed.get('canonical') or '').strip() != 'work_time_entry':
        return None
    mode = str(parsed.get('mode') or '').strip()
    if mode == 'unknown' or mode not in {
        'manual_range',
        'manual_duration',
        'close_at_time',
        'close_with_duration',
        'close_now',
    }:
        return None

    fallback_date = current if open_day is None else date.fromisoformat(open_day.work_date)
    candidate_date = _parse_llm_date(str(parsed.get('date') or ''), fallback=fallback_date)
    if candidate_date is None:
        return None
    start_time = _parse_llm_hhmm(parsed.get('start_time'))
    end_time = _parse_llm_hhmm(parsed.get('end_time'))
    duration_minutes = _parse_llm_duration(parsed.get('duration_minutes'))

    if open_day is not None:
        candidate_date = date.fromisoformat(open_day.work_date)
        if mode == 'close_now':
            return WorkTimeCandidate(work_date=candidate_date, close_mode='close_now', needs_confirmation=False)
        if mode == 'close_with_duration' and duration_minutes is not None:
            return WorkTimeCandidate(work_date=candidate_date, duration_minutes=duration_minutes, close_mode='close_with_duration')
        if mode == 'close_at_time' and end_time is not None:
            return WorkTimeCandidate(work_date=candidate_date, end_time=end_time, close_mode='close_at_time')
        return None

    if mode == 'manual_range' and start_time is not None and end_time is not None:
        return WorkTimeCandidate(work_date=candidate_date, start_time=start_time, end_time=end_time, close_mode='manual_range')
    if mode == 'manual_duration' and duration_minutes is not None:
        return WorkTimeCandidate(work_date=candidate_date, duration_minutes=duration_minutes, close_mode='manual_duration')
    return None



def format_day_summary(day: WorkTimeDay) -> str:
    if day.status == STATUS_CLOSED:
        if day.start_time is None and day.end_time is None:
            return f"{_format_date_sk(day.work_date)}: bez presneho prichodu/odchodu ({_format_duration(day.total_minutes or 0)})"
        return f"{_format_date_sk(day.work_date)}: {day.start_time or '-'} - {day.end_time or '-'} ({_format_duration(day.total_minutes or 0)})"
    if day.status == STATUS_OPEN:
        return f"{_format_date_sk(day.work_date)}: otvoreny od {day.start_time or '-'}"
    if day.status == STATUS_SKIPPED:
        return f"{_format_date_sk(day.work_date)}: preskocene"
    return f"{_format_date_sk(day.work_date)}: {day.status}"


def format_candidate_preview(
    candidate: WorkTimeCandidate,
    *,
    open_day: WorkTimeDay | None = None,
    lunch_break_minutes: int = 0,
) -> str:
    start_value = candidate.start_time
    if start_value is None and open_day is not None and open_day.start_time is not None:
        start_value = _parse_time_value(open_day.start_time)
    end_value = candidate.end_time
    if end_value is None and start_value is not None and candidate.duration_minutes is not None:
        base = datetime.combine(candidate.work_date, start_value)
        end_value = (base + timedelta(minutes=candidate.duration_minutes + max(0, lunch_break_minutes))).time().replace(second=0, microsecond=0)
    elif end_value is None:
        end_value = candidate.calculated_end_time
    total = ''
    if candidate.duration_minutes is not None:
        total = _format_duration(candidate.duration_minutes)
    elif start_value is not None and end_value is not None:
        total_minutes = int(
            (
                datetime.combine(candidate.work_date, end_value)
                - datetime.combine(candidate.work_date, start_value)
            ).total_seconds()
            // 60
        )
        total = _format_duration(max(0, total_minutes - max(0, lunch_break_minutes)))
    entry_type = ''
    if start_value is None and end_value is None and candidate.duration_minutes is not None:
        entry_type = 'Typ: pocet hodin bez presneho prichodu/odchodu\n'
    return (
        f"Datum: {_format_date_sk(candidate.work_date.isoformat())}\n"
        f"{entry_type}"
        f"Prichod: {_format_time(start_value) if start_value else '-'}\n"
        f"Odchod: {_format_time(end_value) if end_value else '-'}\n"
        f"Hodiny: {total or '-'}"
    )


def format_month_summary(summary: WorkTimeMonthSummary) -> str:
    return (
        f"Mesiac: {MONTH_NAMES_SK[summary.month]} {summary.year}\n"
        f"Pocet zaznamov: {summary.row_count}\n"
        f"Spolu hodin: {_format_duration(summary.total_minutes)}"
    )


def _row_to_day(row) -> WorkTimeDay | None:
    if row is None:
        return None
    return WorkTimeDay(
        id=int(row[0]),
        telegram_id=int(row[1]),
        work_date=str(row[2]),
        start_time=row[3],
        end_time=row[4],
        total_minutes=None if row[5] is None else int(row[5]),
        status=str(row[6]),
        source=str(row[7]),
        note=row[8],
        gross_minutes=None if len(row) <= 9 or row[9] is None else int(row[9]),
        lunch_break_minutes_snapshot=None if len(row) <= 10 or row[10] is None else int(row[10]),
        net_work_minutes_override=None if len(row) <= 11 or row[11] is None else int(row[11]),
        close_input_mode=None if len(row) <= 12 or row[12] is None else str(row[12]),
    )


def _effective_lunch_break_minutes(settings: WorkTimeLunchBreakSettings | None) -> int:
    if settings is None or not settings.configured or not settings.enabled:
        return 0
    return max(0, min(180, settings.minutes))


def _validate_lunch_break_minutes(minutes: int) -> int:
    value = int(minutes)
    if value < 0 or value > 180:
        raise ValueError('lunch_break_minutes_out_of_range')
    return value


def _minutes_between(start_value: str, end_value: str) -> int | None:
    try:
        start_time = _parse_time_value(start_value)
        end_time = _parse_time_value(end_value)
    except ValueError:
        return None
    start_dt = datetime.combine(date(2000, 1, 1), start_time)
    end_dt = datetime.combine(date(2000, 1, 1), end_time)
    if end_dt < start_dt:
        return None
    return int((end_dt - start_dt).total_seconds() // 60)


def work_time_timezone_name() -> str:
    return os.getenv(_WORK_TIME_TIMEZONE_ENV, '').strip() or DEFAULT_WORK_TIME_TIMEZONE


def work_time_zoneinfo() -> ZoneInfo:
    timezone_name = work_time_timezone_name()
    try:
        return ZoneInfo(timezone_name)
    except ZoneInfoNotFoundError:
        _LOGGER.warning(
            'Invalid %s=%r; falling back to %s for OfficeFlow work-time clock',
            _WORK_TIME_TIMEZONE_ENV,
            timezone_name,
            DEFAULT_WORK_TIME_TIMEZONE,
        )
        return ZoneInfo(DEFAULT_WORK_TIME_TIMEZONE)


def _runtime_now_utc() -> datetime:
    return datetime.now(UTC)


def work_time_local_now(value: datetime | None = None) -> datetime:
    source = value if value is not None else _runtime_now_utc()
    if source.tzinfo is None:
        return source.replace(second=0, microsecond=0)
    return source.astimezone(work_time_zoneinfo()).replace(second=0, microsecond=0, tzinfo=None)


def work_time_local_date(value: datetime | None = None) -> date:
    return work_time_local_now(value).date()


def _local_now(value: datetime | None = None) -> datetime:
    return work_time_local_now(value)

def _format_time(value: time) -> str:
    return value.replace(second=0, microsecond=0).strftime('%H:%M')


def _parse_time_value(value: str) -> time:
    return datetime.strptime(value, '%H:%M').time()


def _excel_duration_value(total_minutes: int) -> float:
    return int(total_minutes) / (24 * 60)


def _format_duration(total_minutes: int) -> str:
    total_minutes = int(total_minutes)
    hours, minutes = divmod(total_minutes, 60)
    if minutes == 0:
        return f'{hours} hod.'
    decimal_hours = total_minutes / 60
    value = f'{decimal_hours:.2f}'.rstrip('0').rstrip('.').replace('.', ',')
    return f'{value} hod.'


def _parse_duration_minutes(normalized: str) -> int | None:
    match = re.search(
        r'\b(\d{1,3})(?:[,.](\d{1,2}))?\s*(?:hodin|hodiny|hodina|hod|h|godin|casov|chas)\b',
        normalized,
    )
    if not match:
        return None
    hours = int(match.group(1))
    fraction = match.group(2)
    minutes = 0
    if fraction:
        minutes = int(round(float(f'0.{fraction}') * 60))
    total = hours * 60 + minutes
    if total <= 0 or total > 24 * 60:
        return None
    return total


def _format_date_sk(value: str) -> str:
    parsed = date.fromisoformat(value)
    return parsed.strftime('%d.%m.%Y')


def _normalize(value: str) -> str:
    normalized = unicodedata.normalize('NFKD', value.casefold())
    without_marks = ''.join(ch for ch in normalized if not unicodedata.combining(ch))
    return re.sub(r'\s+', ' ', without_marks).strip()


def _parse_llm_date(value: str, *, fallback: date) -> date | None:
    if value in {'', 'None', 'null', 'unknown'}:
        return fallback
    try:
        return date.fromisoformat(value)
    except ValueError:
        return None


def _parse_llm_hhmm(value: object) -> time | None:
    if value is None:
        return None
    raw = str(value).strip()
    if raw.lower() in {'', 'none', 'null', 'unknown'}:
        return None
    if not re.fullmatch(r'\d{2}:\d{2}', raw):
        return None
    try:
        return _parse_time_value(raw)
    except ValueError:
        return None


def _parse_llm_duration(value: object) -> int | None:
    if value is None:
        return None
    try:
        total = int(value)
    except (TypeError, ValueError):
        return None
    if total <= 0 or total > 24 * 60:
        return None
    return total


def _parse_match_time(match: tuple[str, str]) -> time | None:
    hour = int(match[0])
    minute = int(match[1] or '0')
    if hour > 23 or minute > 59:
        return None
    return time(hour=hour, minute=minute)


def _resolve_work_date_and_strip_explicit_date(normalized: str, *, today: date | None = None) -> tuple[date, str]:
    current = today or work_time_local_date()
    explicit = _parse_explicit_work_date(normalized, today=current)
    if explicit is not None:
        work_date, span = explicit
        return work_date, f'{normalized[:span[0]]} {normalized[span[1]:]}'
    return _resolve_relative_date(normalized, today=current), normalized


def _parse_explicit_work_date(normalized: str, *, today: date) -> tuple[date, tuple[int, int]] | None:
    patterns = (
        r'\b(?P<day>0?[1-9]|[12]\d|3[01])\s*[./-]\s*(?P<month>0?[1-9]|1[0-2])(?:\s*[./-]\s*(?P<year>(?:19|20)\d{2}))?\b',
        r'\b(?P<day>0?[1-9]|[12]\d|3[01])\s+(?:na|\u0437\u0430|\u043d\u0430)\s+(?P<month>0?[1-9]|1[0-2])(?:\s+(?P<year>(?:19|20)\d{2}))?\b',
    )
    for pattern in patterns:
        match = re.search(pattern, normalized)
        if not match:
            continue
        day = int(match.group('day'))
        month = int(match.group('month'))
        year = int(match.group('year') or today.year)
        try:
            return date(year, month, day), match.span()
        except ValueError:
            return None
    return None

def _resolve_relative_date(normalized: str, *, today: date | None = None) -> date:
    current = today or work_time_local_date()
    yesterday_terms = {
        'vcera',
        'vchera',
        'вчора',
        'учора',
        'вчера',
        'вчорашній',
        'вчерашний',
    }
    if any(term in normalized for term in yesterday_terms):
        return current - timedelta(days=1)
    return current


_MONTH_ALIASES = {
    1: ('januar', 'jan', 'СЃС–С‡РµРЅСЊ', 'СЏРЅРІР°СЂ'),
    2: ('februar', 'feb', 'Р»СЋС‚РёР№', 'С„РµРІСЂР°Р»'),
    3: ('marec', 'marci', 'march', 'Р±РµСЂРµР·РµРЅСЊ', 'РјР°СЂС‚'),
    4: ('april', 'apr', 'РєРІС–С‚РµРЅСЊ', 'Р°РїСЂРµР»'),
    5: ('maj', 'may', 'С‚СЂР°РІРµРЅСЊ', 'РјР°Р№'),
    6: ('jun', 'june', 'jún', 'С‡РµСЂРІРµРЅСЊ', 'РёСЋРЅ'),
    7: ('jul', 'july', 'júl', 'Р»РёРїРµРЅСЊ', 'РёСЋР»'),
    8: ('august', 'aug', 'СЃРµСЂРїРµРЅСЊ', 'Р°РІРіСѓСЃС‚'),
    9: ('september', 'sep', 'РІРµСЂРµСЃРµРЅСЊ', 'СЃРµРЅС‚СЏР±СЂ'),
    10: ('október', 'oct', 'Р¶РѕРІС‚РµРЅСЊ', 'РѕРєС‚СЏР±СЂ'),
    11: ('november', 'nov', 'Р»РёСЃС‚РѕРїР°Рґ', 'РЅРѕСЏР±СЂ'),
    12: ('december', 'dec', 'РіСЂСѓРґРµРЅСЊ', 'РґРµРєР°Р±СЂ'),
}
